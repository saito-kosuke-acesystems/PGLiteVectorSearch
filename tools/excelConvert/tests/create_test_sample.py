"""
図形付きExcelファイルのサンプル作成スクリプト
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def create_sample_excel_with_shapes():
    """図形付きサンプルExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Sheet"
    
    # 基本データを追加
    ws['A1'] = "図形テスト用サンプル"
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A3'] = "名前"
    ws['B3'] = "値"
    ws['A4'] = "データ1"
    ws['B4'] = 100
    ws['A5'] = "データ2"
    ws['B5'] = 200
    ws['A6'] = "データ3"
    ws['B6'] = 150
    
    # セルのスタイルを設定
    for row in range(3, 7):
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            if row == 3:  # ヘッダー行
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ファイルを保存
    sample_file = "sample_with_shapes.xlsx"
    wb.save(sample_file)
    print(f"サンプルファイルを作成しました: {sample_file}")
    
    return sample_file

def test_with_sample():
    """作成したサンプルでテスト"""
    # プロジェクトルートをPythonパスに追加
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(current_dir))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)
    
    from tools.excelConvert.logics.excelToHtmlLogic import excel_to_html
    
    print("=== サンプルExcelファイルでテスト ===")
    
    # サンプルExcelファイルを作成
    excel_file = create_sample_excel_with_shapes()
    output_file = "sample_output.html"
    
    try:
        print(f"入力ファイル: {excel_file}")
        print(f"出力ファイル: {output_file}")
        
        # 変換実行
        excel_to_html(excel_file, output_file)
        
        print(f"変換完了: {output_file}")
        
        # ブラウザで開く
        try:
            os.startfile(output_file)
        except:
            print("HTMLファイルを手動で開いてください。")
            
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_with_sample()
