"""
図形付きExcelファイルの簡単作成ツール
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import tempfile

def create_simple_excel_with_basic_shapes():
    """基本的な図形付きExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "図形テスト"
    
    # 基本データを追加
    ws['A1'] = "図形テストファイル"
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A3'] = "処理フロー"
    ws['A3'].font = Font(bold=True, size=14)
    
    # テーブルデータ
    headers = ["内容", "サンプルプロジェクト", "成果物名", "システム担当者名", "作業", "ITS", "2019-03-10"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # データ行
    data_rows = [
        ["システム名", "サンプルシステム", "", "", "変更", "ITS", "2022-12-19"],
        ["サブシステム名", "顧客管理システム", "", "", "", "", ""]
    ]
    
    for row_idx, row_data in enumerate(data_rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    try:
        # 簡単な画像を作成（1x1ピクセルのPNG）
        from openpyxl.drawing.image import Image
        
        # 1x1の透明PNGデータ（Base64エンコード済み）
        png_data = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x01\x00\x00\x00\x007n\xf9$\x00\x00\x00\nIDAT\x08\x1dc\xf8\x00\x00\x00\x01\x00\x01um!(\x00\x00\x00\x00IEND\xaeB`\x82'
        
        # 画像をファイルとして一時保存
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
            tmp_file.write(png_data)
            tmp_filename = tmp_file.name
        
        # 画像をワークシートに追加
        img = Image(tmp_filename)
        img.width = 100
        img.height = 50
        ws.add_image(img, 'B8')  # B8セルに配置
        
        # 一時ファイルを削除
        os.unlink(tmp_filename)
        
        print("画像を追加しました")
        
    except Exception as e:
        print(f"画像追加でエラー: {e}")
    
    # ファイルを保存
    filename = "test_shapes_simple.xlsx"
    wb.save(filename)
    print(f"図形付きExcelファイルを作成しました: {filename}")
    
    return filename

def test_simple_shapes():
    """作成した図形付きExcelファイルでテスト"""
    # プロジェクトルートをPythonパスに追加
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(current_dir))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)
    
    from tools.excelConvert.logics.excelToHtmlLogic import excel_to_html
    
    print("=== 簡単な図形付きExcelファイルでテスト ===")
    
    # 図形付きExcelファイルを作成
    excel_file = create_simple_excel_with_basic_shapes()
    output_file = "simple_shapes_output.html"
    
    try:
        print(f"入力ファイル: {excel_file}")
        print(f"出力ファイル: {output_file}")
        
        # 変換実行
        excel_to_html(excel_file, output_file)
        
        print(f"変換完了: {output_file}")
        
        # HTMLファイルの内容をチェック
        with open(output_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        if 'excel-shape' in html_content:
            print("✓ HTMLに図形クラスが含まれています")
        else:
            print("✗ HTMLに図形クラスが含まれていません")
            
        if 'data:image' in html_content:
            print("✓ HTMLに画像データが含まれています")
        else:
            print("✗ HTMLに画像データが含まれていません")
        
        # ブラウザで開く
        try:
            os.startfile(output_file)
        except:
            print("HTMLファイルを手動で開いてください。")
            
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_simple_shapes()
