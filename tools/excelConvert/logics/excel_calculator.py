"""
Excel 計算関連のユーティリティ関数
"""


def get_formula_cells_from_openpyxl(excel_file: str):
    """openpyxlを使用して関数セルの位置を事前に特定"""
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(excel_file, data_only=False)
        formula_cells = {}
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            formula_cells[sheet_name] = []
            
            # 使用範囲内の全セルをチェック
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # 関数セル
                        formula_cells[sheet_name].append({
                            'row': cell.row,
                            'col': cell.column,
                            'formula': cell.value
                        })
                        print(f"関数セル発見: {sheet_name} 行{cell.row}, 列{cell.column} = {cell.value}")
        
        return formula_cells
        
    except Exception as e:
        print(f"openpyxlでの関数セル特定に失敗: {e}")
        return {}


def get_excel_calculated_values_focused(excel_file: str, formula_cells):
    """特定された関数セルに焦点を当てて計算結果を取得"""
    try:
        import xlwings as xw
        
        print("関数セルに焦点を当て計算結果を取得中...")
        
        # 非表示でExcelアプリケーションを開始
        app = xw.App(visible=False, add_book=False)
        
        try:
            # ワークブックを開く
            wb = app.books.open(excel_file)
            
            # 計算方式を自動に設定
            app.calculation = 'automatic'
            
            # 計算を強制実行
            wb.app.calculate()
            wb.app.calculate()
            
            calculated_values = {}
            
            for sheet_name, formulas in formula_cells.items():
                if not formulas:
                    continue
                    
                calculated_values[sheet_name] = {}
                sheet = wb.sheets[sheet_name]
                
                for formula_info in formulas:
                    row = formula_info['row']
                    col = formula_info['col']
                    formula = formula_info['formula']
                    
                    try:
                        # 個別に関数セルの値を取得
                        cell = sheet.range((row, col))
                        value = cell.value
                        
                        calculated_values[sheet_name][(row, col)] = value
                        print(f"✓ 関数計算成功: {sheet_name} 行{row}, 列{col}: {formula} → {value}")
                        
                    except Exception as e:
                        print(f"⚠ 関数計算失敗: {sheet_name} 行{row}, 列{col}: {formula} → エラー: {e}")
            
            # ワークブックを閉じる
            wb.close()
            
            return calculated_values
            
        finally:
            # アプリケーションを確実に終了
            try:
                app.quit()
            except:
                pass
                
    except Exception as e:
        print(f"関数セル特化計算でエラー: {e}")
        return {}


def get_excel_calculated_values_xlwings(excel_file: str):
    """xlwingsを使用してExcelファイルから関数の計算結果を取得"""
    try:
        import xlwings as xw
        
        print("xlwingsでExcelファイルを開いて計算実行中...")
        
        # 非表示でExcelアプリケーションを開始
        app = xw.App(visible=False, add_book=False)
        
        try:
            # ワークブックを開く
            wb = app.books.open(excel_file)
            
            # 計算方式を自動に設定
            app.calculation = 'automatic'
            
            # 計算を強制実行（複数回実行して確実に計算）
            wb.app.calculate()
            wb.app.calculate()  # 2回実行して確実に
            
            # 計算結果を辞書に保存
            calculated_values = {}
            
            for sheet in wb.sheets:
                sheet_name = sheet.name
                calculated_values[sheet_name] = {}
                
                # 使用範囲を取得
                used_range = sheet.used_range
                if used_range:
                    print(f"シート '{sheet_name}': 使用範囲 {used_range.address}")
                    
                    # 個別にセルの値を取得（より確実な方法）
                    start_row = used_range.row
                    start_col = used_range.column
                    end_row = start_row + used_range.rows.count - 1
                    end_col = start_col + used_range.columns.count - 1
                    
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            try:
                                # 個別のセルにアクセス
                                cell = sheet.range((row, col))
                                value = cell.value
                                
                                if value is not None:
                                    calculated_values[sheet_name][(row, col)] = value
                                    
                                    # デバッグ: 関数セルかどうかをチェック
                                    try:
                                        formula = cell.formula
                                        if formula and formula.startswith('='):
                                            print(f"関数セル {sheet_name} 行{row}, 列{col}: {formula} → {value}")
                                    except:
                                        pass  # 関数でない場合は無視
                                        
                            except Exception as e:
                                # 個別セルの取得に失敗した場合は無視
                                print(f"セル {row},{col} の取得に失敗: {e}")
                                continue
                    
                    print(f"シート '{sheet_name}': {len(calculated_values[sheet_name])}個のセル値を取得")
            
            # ワークブックを閉じる
            wb.close()
            
        finally:
            # アプリケーションを確実に終了
            try:
                app.quit()
            except:
                pass
        
        # 取得できた値があるかチェック
        total_values = sum(len(sheet_values) for sheet_values in calculated_values.values())
        print(f"xlwingsで合計{total_values}個のセル値を取得しました")
        
        return calculated_values if total_values > 0 else None
        
    except ImportError:
        print("xlwingsがインストールされていません。")
        return None
    except Exception as e:
        print(f"Excel計算値取得エラー (xlwings): {e}")
        return None
