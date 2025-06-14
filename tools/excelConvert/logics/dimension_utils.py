"""
Excel 寸法計算関連のユーティリティ関数
"""
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def excel_to_pixels(excel_width_units: float, is_width: bool = True) -> int:
    """Excelの幅・高さの単位をピクセルに変換
    
    Args:
        excel_width_units: Excelの幅・高さの値
        is_width: Trueの場合は列幅、Falseの場合は行高
    
    Returns:
        ピクセル値
    """
    if is_width:
        # 列幅の変換: より正確な係数を使用
        # Excelの列幅は文字数ベースで、1単位 ≈ 6ピクセル程度が実際に近い
        return max(int(excel_width_units * 6.0), 1)  # より小さい係数で正確に
    else:
        # 行高の変換: 1ポイント = 1.33ピクセル
        return max(int(excel_width_units * 1.33), 20)  # 最小高20px


def get_default_dimensions(worksheet) -> tuple:
    """ワークシートのデフォルト列幅と行高を取得
    
    Returns:
        (default_column_width_px, default_row_height_px)
    """
    # デフォルト列幅（Excelの実際のデフォルト値を使用）
    default_col_width = getattr(worksheet, 'default_column_width', 8.43)  # Excel本来のデフォルト
    default_col_width_px = excel_to_pixels(default_col_width, True)
    
    # デフォルト行高（Excelのデフォルトは15ポイント）
    default_row_height = getattr(worksheet, 'default_row_height', 15)
    default_row_height_px = excel_to_pixels(default_row_height, False)
    
    print(f"デフォルト寸法: 列幅 {default_col_width} → {default_col_width_px}px, 行高 {default_row_height} → {default_row_height_px}px")
    
    return default_col_width_px, default_row_height_px


def calculate_optimal_column_widths(worksheet, max_row, max_col) -> dict:
    """セル内容に基づいて最適な列幅を計算
    
    Args:
        worksheet: openpyxlワークシート
        max_row: 最大行数
        max_col: 最大列数
    
    Returns:
        列番号をキーとした幅のピクセル値の辞書
    """
    column_widths = {}
    
    for col_idx in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)  # 安全な方法で列文字を取得
        
        # 各行の該当列のセル内容の長さを計算
        for row_idx in range(1, max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            # MergedCellの場合はスキップ
            if isinstance(cell, MergedCell):
                continue
                
            if cell.value is not None:
                # セルの値を文字列に変換して長さを計算
                cell_text = str(cell.value)
                
                # 日本語文字の場合は1.5倍の幅を考慮
                japanese_char_count = sum(1 for char in cell_text if ord(char) > 127)
                english_char_count = len(cell_text) - japanese_char_count
                effective_length = english_char_count + (japanese_char_count * 1.5)
                
                max_length = max(max_length, effective_length)        # 最小幅と最大幅を設定（Excelの実際の設定を尊重）
        min_width = 8    # 最小8ピクセル（Excelの極小列に対応）
        max_width = 400  # 最大400ピクセル
        
        # 文字数に基づく幅計算（より控えめに）
        if max_length == 0:
            optimal_width = min_width
        else:
            optimal_width = int(max_length * 6 + 10)  # 文字幅6px + パディング10px（控えめに）
            optimal_width = max(min_width, min(optimal_width, max_width))
        
        column_widths[col_idx] = optimal_width
    
    return column_widths


def calculate_optimal_row_heights(worksheet, max_row, max_col) -> dict:
    """セル内容に基づいて最適な行高を計算
    
    Args:
        worksheet: openpyxlワークシート
        max_row: 最大行数
        max_col: 最大列数
    
    Returns:
        行番号をキーとした高さのピクセル値の辞書
    """
    row_heights = {}
    
    for row_idx in range(1, max_row + 1):
        max_lines = 1
        
        # 行内の全セルで改行数をチェック
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_text = str(cell.value)
                lines = cell_text.count('\n') + 1
                max_lines = max(max_lines, lines)
        
        # 行高を計算（基本高さ25px + 追加行ごとに20px）
        base_height = 25
        additional_height = (max_lines - 1) * 20
        row_heights[row_idx] = base_height + additional_height
    
    return row_heights


def get_cell_dimensions(worksheet) -> tuple:
    """ワークシートの使用範囲を取得"""
    if worksheet.max_row == 1 and worksheet.max_column == 1:
        # 空のシートの場合
        return 0, 0
    return worksheet.max_row, worksheet.max_column
