from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import html
import os
import sys
from typing import Optional

# プロジェクトルートをPythonパスに追加
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_dir)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# 分割されたモジュールをインポート
from tools.excelConvert.logics.style_utils import get_cell_style
from tools.excelConvert.logics.dimension_utils import (
    get_cell_dimensions, 
    get_default_dimensions,
    calculate_optimal_column_widths,
    calculate_optimal_row_heights,
    excel_to_pixels
)
from tools.excelConvert.logics.excel_calculator import (
    get_excel_calculated_values_xlwings,
    get_formula_cells_from_openpyxl,
    get_excel_calculated_values_focused
)
from tools.excelConvert.logics.merge_utils import (
    get_merged_cell_info,
    should_skip_cell,
    get_cell_merge_attributes,
    get_merged_cell_value
)
from tools.excelConvert.logics.html_generator import (
    generate_html_header,
    generate_html_footer,
    generate_sheet_header,
    generate_sheet_footer,
    generate_colgroup,
    generate_table_with_width,
    generate_shapes_html
)
from tools.excelConvert.logics.shape_utils import get_sheet_shapes, get_sheet_shapes_alternative, get_sheet_shapes_deep_inspection


def excel_to_html(excel_file: str, output_file: str, sheet_names: Optional[list] = None):
    """
    Excelファイルを外観を保ったHTMLファイルに変換
    
    Args:
        excel_file (str): 入力Excelファイルパス
        output_file (str): 出力HTMLファイルパス
        sheet_names (list, optional): 変換するシート名のリスト。Noneの場合は全シート
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"ファイルが見つかりません: {excel_file}")
      # Excel関数の計算結果を取得（複数の方法を試行）
    print("Excel関数の計算結果を取得中...")
    calculated_values = None
    
    # 方法1: 事前に関数セルを特定してから計算
    print("方法1: 関数セルを特定してから計算...")
    try:
        formula_cells = get_formula_cells_from_openpyxl(excel_file)
        if any(formulas for formulas in formula_cells.values()):
            calculated_values = get_excel_calculated_values_focused(excel_file, formula_cells)
            if calculated_values:
                total_values = sum(len(sheet_values) for sheet_values in calculated_values.values())
                print(f"方法1で{total_values}個の関数セル値を取得しました")
    except Exception as e:
        print(f"方法1での取得に失敗: {e}")
    
    # 方法2: 従来のxlwings方式（方法1が失敗した場合）
    if not calculated_values or sum(len(sheet_values) for sheet_values in calculated_values.values()) == 0:
        print("方法2: xlwingsで計算結果を取得中...")
        try:
            calculated_values = get_excel_calculated_values_xlwings(excel_file)
        except Exception as e:
            print(f"方法2での取得に失敗: {e}")
    
    # 取得に失敗した場合は空の辞書を使用
    if calculated_values is None:
        print("関数の計算結果取得に失敗。フォーミュラを表示します。")
        calculated_values = {}

    # Excelファイルを読み込み（スタイル情報取得のため）
    workbook = load_workbook(excel_file, data_only=False)
    
    # 変換対象のシートを決定
    if sheet_names is None:
        target_sheets = workbook.sheetnames
    else:
        target_sheets = [name for name in sheet_names if name in workbook.sheetnames]
    
    if not target_sheets:
        raise ValueError("変換対象のシートが見つかりません")
    
    # HTMLの開始
    html_content = generate_html_header(excel_file)
      # 各シートを処理
    for sheet_name in target_sheets:
        worksheet = workbook[sheet_name]
        max_row, max_col = get_cell_dimensions(worksheet)
        
        if max_row == 0 or max_col == 0:
            # 空のシートの場合はスキップ
            continue
        
        # 結合セル情報を取得
        merged_info = get_merged_cell_info(worksheet)
        merged_cells_map = merged_info['merged_cells_map']
        print(f"シート '{sheet_name}': {len(merged_info['merged_ranges'])}個の結合セル範囲を検出")        # 図形情報を取得
        print(f"シート '{sheet_name}': 図形を取得中...")
        shapes = get_sheet_shapes(worksheet)
        
        # 標準方法で図形が見つからない場合は代替方法を試す
        if not shapes:
            print(f"標準方法で図形が見つからないため、代替方法を試行中...")
            shapes_alt = get_sheet_shapes_alternative(worksheet)
            shapes.extend(shapes_alt)
        
        # それでも見つからない場合は深い検査を実行
        if not shapes:
            print(f"深いレベルでの図形検査を実行中...")
            shapes_deep = get_sheet_shapes_deep_inspection(worksheet)
            shapes.extend(shapes_deep)
        
        print(f"シート '{sheet_name}': {len(shapes)}個の図形を検出")
        
        # シートヘッダーを追加
        html_content += generate_sheet_header(sheet_name)
          # 列幅を設定するためのcolgroup要素を追加
        default_col_width_px, default_row_height_px = get_default_dimensions(worksheet)
        
        # セル内容に基づく最適な寸法を計算
        optimal_column_widths = calculate_optimal_column_widths(worksheet, max_row, max_col)
        optimal_row_heights = calculate_optimal_row_heights(worksheet, max_row, max_col)
        
        print(f"シート '{sheet_name}': {max_row}行 × {max_col}列")
        print(f"デフォルト列幅: {default_col_width_px}px, デフォルト行高: {default_row_height_px}px")        # 全列の寸法情報をデバッグ出力
        print("=== 列寸法情報 ===")
        print(f"ワークシートのdefault_column_width: {getattr(worksheet, 'default_column_width', 'なし')}")
        for col_idx in range(1, min(max_col + 1, 10)):  # 最初の10列まで表示
            column_letter = get_column_letter(col_idx)
            col_dim = worksheet.column_dimensions.get(column_letter)
            if col_dim:
                print(f"  列{column_letter}: width={getattr(col_dim, 'width', 'なし')}, auto={getattr(col_dim, 'auto', 'なし')}, bestFit={getattr(col_dim, 'bestFit', 'なし')}")
            else:
                print(f"  列{column_letter}: column_dimensionsに設定なし")
        
        # テーブル全体の幅を計算
        total_table_width = 0
        column_widths_for_table = []
        
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)  # 安全な方法で列文字を取得
            
            # 明示的に設定された列幅があるかチェック（より厳密に）
            col_dim = worksheet.column_dimensions.get(column_letter)
            if col_dim and hasattr(col_dim, 'width') and col_dim.width is not None and col_dim.width > 0:
                col_width = col_dim.width
                width_px = excel_to_pixels(col_width, True)
                print(f"列 {column_letter}: Excelで設定された幅 {col_width} → {width_px}px")
            else:
                # セル内容に基づく最適な幅を使用（より控えめに）
                optimal_width = optimal_column_widths.get(col_idx, default_col_width_px)
                # 最適幅も元のExcelに近づけるため、さらに小さく調整
                width_px = min(optimal_width, default_col_width_px * 1.2)  # デフォルトの1.2倍以下に制限
                print(f"列 {column_letter}: 計算された最適な幅 → {width_px}px (制限適用)")
            
            # Excelの実際の設定をそのまま使用（最小幅制限なし）
            print(f"列 {column_letter}: 最終適用幅 → {width_px}px")
            
            column_widths_for_table.append(width_px)
            total_table_width += width_px
        
        # テーブルスタイルを更新（固定幅に設定）
        html_content = html_content.replace('<table>', generate_table_with_width(total_table_width))
        
        # colgroupを追加
        html_content += generate_colgroup(column_widths_for_table)
        
        # セルデータを処理
        for row_idx in range(1, max_row + 1):
            # 行の高さを取得してスタイルに追加
            row_style = ""
            # 明示的に設定された行高があるかチェック
            if row_idx in worksheet.row_dimensions and worksheet.row_dimensions[row_idx].height:
                row_height = worksheet.row_dimensions[row_idx].height
                height_px = excel_to_pixels(row_height, False)
                row_style = f' style="height: {height_px}px;"'
            else:
                # セル内容に基づく最適な高さを使用
                optimal_height = optimal_row_heights.get(row_idx, default_row_height_px)
                row_style = f' style="height: {optimal_height}px;"'
            
            html_content += f"            <tr{row_style}>\n"
            
            for col_idx in range(1, max_col + 1):
                # 結合セルでスキップ対象の場合は処理しない
                if should_skip_cell(row_idx, col_idx, merged_cells_map):
                    continue
                
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # MergedCellの場合の安全なログ出力
                if row_idx <= 3 and col_idx <= 3:
                    if isinstance(cell, MergedCell):
                        print(f"結合セル {row_idx},{col_idx}: MergedCell")
                    else:
                        print(f"セル {row_idx},{col_idx}: フォント={cell.font}, 塗りつぶし={cell.fill}")
                
                # 結合セルの値を適切に取得
                cell_value = get_merged_cell_value(
                    worksheet, row_idx, col_idx, merged_cells_map, 
                    calculated_values, sheet_name
                )
                
                # セルのスタイルを取得（幅と高さ情報を含む）
                cell_style = get_cell_style(cell, worksheet, row_idx, col_idx)
                
                # セルのスタイルが空の場合のデフォルト設定
                if not cell_style:
                    cell_style = "background-color: #FFFFFF; color: #000000; font-family: 'Calibri', Arial, sans-serif; font-size: 11pt"
                
                # 空のセルの場合はクラスを追加、改行が含まれる場合もクラスを追加
                css_classes = []
                if not cell_value:
                    css_classes.append("empty-cell")
                if '\n' in str(cell_value):
                    css_classes.append("multiline")
                
                css_class = f' class="{" ".join(css_classes)}"' if css_classes else ''
                
                # 結合セルの属性を取得
                merge_attrs = get_cell_merge_attributes(row_idx, col_idx, merged_cells_map)
                
                # HTMLに追加（結合セル属性を含める）
                html_content += f'                <td style="{cell_style}"{css_class}{merge_attrs}>{html.escape(cell_value)}</td>\n'
            
            html_content += "            </tr>\n"
          # テーブル終了
        html_content += "        </table>\n"
        
        # 図形をHTMLに追加（テーブルの後、sheet-containerの終了前に）
        if shapes:
            print(f"図形を{len(shapes)}個HTMLに追加中...")
            
            # 行の高さリストを作成
            row_heights_list = []
            for row_idx in range(1, max_row + 1):
                if row_idx in worksheet.row_dimensions and worksheet.row_dimensions[row_idx].height:
                    row_height = worksheet.row_dimensions[row_idx].height
                    height_px = excel_to_pixels(row_height, False)
                    row_heights_list.append(height_px)
                else:
                    optimal_height = optimal_row_heights.get(row_idx, default_row_height_px)
                    row_heights_list.append(optimal_height)
            
            # 図形のHTMLを生成
            shapes_html = generate_shapes_html(shapes, column_widths_for_table, row_heights_list)
            html_content += shapes_html
            print(f"図形HTML追加完了")
        
        # sheet-container終了
        html_content += "    </div>\n"
    
    # HTMLの終了
    html_content += generate_html_footer()
    
    # HTMLファイルに保存
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"HTMLファイルが作成されました: {output_file}")
    print(f"変換されたシート数: {len(target_sheets)}")
    for sheet_name in target_sheets:
        print(f"  - {sheet_name}")


def main():
    """メイン実行関数"""
    # 使用例
    excel_file = "sample.xlsx"  # 変換したいExcelファイルのパス
    output_file = "output.html"  # 出力HTMLファイルのパス
    
    try:
        # 全シートを変換
        excel_to_html(excel_file, output_file)
        
        # 特定のシートのみ変換する場合
        # excel_to_html(excel_file, output_file, sheet_names=["Sheet1", "Sheet2"])
        
    except FileNotFoundError as e:
        print(f"エラー: {e}")
        print("sample.xlsxファイルを同じディレクトリに配置してください。")
    except Exception as e:
        print(f"変換中にエラーが発生しました: {e}")


if __name__ == "__main__":
    main()
