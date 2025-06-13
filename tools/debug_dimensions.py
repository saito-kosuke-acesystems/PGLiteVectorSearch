#!/usr/bin/env python3
"""
Excelファイルの寸法情報をデバッグするスクリプト
"""

from openpyxl import load_workbook
import os

def debug_excel_dimensions(excel_file):
    """Excelファイルの寸法情報を詳細に表示"""
    print(f"=== {excel_file} の寸法情報 ===\n")
    
    if not os.path.exists(excel_file):
        print(f"ファイルが見つかりません: {excel_file}")
        return
    
    workbook = load_workbook(excel_file)
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        print(f"シート: {sheet_name}")
        print("-" * 50)
        
        # デフォルト寸法
        print(f"デフォルト列幅: {getattr(worksheet, 'default_column_width', 'なし')}")
        print(f"デフォルト行高: {getattr(worksheet, 'default_row_height', 'なし')}")
        print()
        
        # 列幅情報
        print("列幅情報:")
        if worksheet.column_dimensions:
            for col, dim in worksheet.column_dimensions.items():
                print(f"  列 {col}: 幅={dim.width}, 非表示={dim.hidden}")
        else:
            print("  設定された列幅なし")
        print()
        
        # 行高情報
        print("行高情報:")
        if worksheet.row_dimensions:
            for row, dim in worksheet.row_dimensions.items():
                print(f"  行 {row}: 高さ={dim.height}, 非表示={dim.hidden}")
        else:
            print("  設定された行高なし")
        print()
        
        # 使用範囲
        print(f"使用範囲: 行1-{worksheet.max_row}, 列1-{worksheet.max_column}")
        print(f"使用範囲アドレス: {worksheet.calculate_dimension()}")
        print()
        
        # 実際のセル幅を推測（文字数ベース）
        print("セル内容と推奨幅:")
        for row_idx in range(1, min(worksheet.max_row + 1, 6)):  # 最初の5行のみ
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    content_length = len(str(cell.value))
                    suggested_width = max(content_length * 1.2, 8)  # 文字数 * 1.2、最小8
                    print(f"  {cell.coordinate}: '{cell.value}' (長さ:{content_length}, 推奨幅:{suggested_width:.1f})")
        print("\n" + "=" * 60 + "\n")

if __name__ == "__main__":
    debug_excel_dimensions("sample.xlsx")
