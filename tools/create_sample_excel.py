import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_excel():
    """サンプルのExcelファイルを作成"""
    wb = openpyxl.Workbook()
    
    # 最初のシート
    ws1 = wb.active
    ws1.title = "売上データ"
    
    # ヘッダー行
    headers = ["商品名", "価格", "売上数", "合計"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
      # データ行
    data = [
        ["商品A", 1000, 50, "=B2*C2"],
        ["商品B", 1500, 30, "=B3*C3"],
        ["商品C", 800, 70, "=B4*C4"],
        ["合計", "=SUM(B2:B4)", "=SUM(C2:C4)", "=SUM(D2:D4)"]
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 5:  # 合計行
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # 2番目のシート
    ws2 = wb.create_sheet("顧客情報")
    
    # ヘッダー
    headers2 = ["顧客ID", "会社名", "担当者", "電話番号"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # データ
    customer_data = [
        ["C001", "株式会社ABC", "田中太郎", "03-1234-5678"],
        ["C002", "XYZ商事", "佐藤花子", "06-9876-5432"],
        ["C003", "テスト株式会社", "山田次郎", "052-1111-2222"]
    ]
    
    for row_idx, row_data in enumerate(customer_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # 3番目のシート（空のシート）
    ws3 = wb.create_sheet("空のシート")
    
    # ファイル保存
    wb.save("sample.xlsx")
    print("サンプルExcelファイル 'sample.xlsx' を作成しました")

if __name__ == "__main__":
    create_sample_excel()
